"""
POST /api/detect — 통합 비식별화 엔드포인트.

파일 확장자를 보고 내부적으로 5종 detector로 분기한다.
PHP 팀은 파일 형식에 관계없이 이 단일 엔드포인트만 사용한다.

분기 정책:
  .xlsx → xlsx apply (applyMode="applied", downloadToken 생성)
  .docx → docx guide
  .pptx → pptx guide
  .hwpx → hwpx guide
  .pdf  → pdf guide

응답 구조 (C안):
  success, metadata를 최상위에 추가하되 result 중첩 없음.
  PHP에서 $response['autoResults']로 직접 접근 가능.

  성공:
    {
      "success": true,
      "fileType": "...",
      "applyMode": "...",
      "downloadToken": null | "abc123",
      "outputFilePath": null,
      "autoResults": [...],
      "reviewTargets": [...],
      "warnings": [...],
      "summary": {...},
      "metadata": {
        "originalFilename": "test.pdf",
        "fileSize": 12345
      }
    }

  에러:
    {
      "success": false,
      "errorCode": "UNSUPPORTED_FILE_TYPE",
      "message": "지원하지 않는 파일 형식입니다.",
      "detail": {"filename": "test.txt"}
    }

모델 경로:
  NER_MODEL_PATH  환경 변수로 지정. 미설정 시 NER 비활성.
  AI_MODEL_PATH   환경 변수로 지정. 미설정 시 AI 비활성.
  useNer/useAi=true인데 모델 미설정 시 HTTP 503 반환.

downloadToken 저장소:
  프로세스 메모리 기반 — 단일 워커에서만 보장됨.
  멀티 워커 배포 시 Redis/SQLite/DB 기반으로 교체 필요.
  운영 실행: uvicorn ... --workers 1
"""

from __future__ import annotations

import os
import secrets
import sys
from pathlib import Path
from typing import Any

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse
from starlette.background import BackgroundTask

SRC_DIR = Path(__file__).resolve().parent.parent
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from api.errors import detect_file_type
from api.files import (
    _token_store_remove,
    cleanup_expired_tokens,
    cleanup_orphan_files,
    create_download_token,
    get_tmp_dir,
    resolve_download_token,
    revoke_token,
    save_upload_tmp,
)

router = APIRouter()

# ── 기본값 ────────────────────────────────────────────────────

DEFAULT_DELETION_MODE = "mark"   # 검토 피드백 반영: delete → mark
DEFAULT_USE_NER = False
DEFAULT_USE_AI = False


# ── 에러 코드 ──────────────────────────────────────────────────

def _error(status: int, code: str, message: str, **detail) -> JSONResponse:
    """통일된 에러 응답 생성."""
    body: dict[str, Any] = {
        "success": False,
        "errorCode": code,
        "message": message,
    }
    if detail:
        body["detail"] = detail
    return JSONResponse(status_code=status, content=body)


def _success(result_dict: dict[str, Any], filename: str, filesize: int) -> JSONResponse:
    """통일된 성공 응답 생성 (C안: success + metadata 추가, result 중첩 없음)."""
    body = {
        "success": True,
        **result_dict,
        "metadata": {
            "originalFilename": filename,
            "fileSize": filesize,
        },
    }
    return JSONResponse(content=body)


# ── 모델 로더 ──────────────────────────────────────────────────

def _get_ner_model_path() -> str:
    """환경 변수 NER_MODEL_PATH에서 모델 경로를 읽는다."""
    return os.environ.get("NER_MODEL_PATH", "").strip()


def _get_ai_model_path() -> str:
    """환경 변수 AI_MODEL_PATH에서 모델 경로를 읽는다."""
    return os.environ.get(
        "AI_MODEL_PATH",
        "models/privacy_cso_char_keras_model.keras",
    ).strip()


def _get_ner_threshold() -> float:
    """NER 탐지 신뢰도 임계값.

    환경 변수 NER_THRESHOLD로 조절. 기본 0.8.
    값이 낮을수록 더 많이 탐지 (오탐 증가).
    값이 높을수록 더 적게 탐지 (누락 증가).
    """
    try:
        return float(os.environ.get("NER_THRESHOLD", "0.8"))
    except ValueError:
        return 0.8


def _get_ai_threshold() -> float:
    """AI 분류 신뢰도 임계값.

    환경 변수 AI_THRESHOLD로 조절. 기본 0.5.
    값이 낮을수록 더 많이 review_targets에 포함.
    값이 높을수록 더 적게 포함.

    참고: 모델 예측값이 0.4~0.6 사이에 분포하는 경우가 많으므로
    기본값을 0.6에서 0.5로 낮춤 (기존 detect_router 0.6보다 낮게 설정).
    """
    try:
        return float(os.environ.get("AI_THRESHOLD", "0.5"))
    except ValueError:
        return 0.5


def get_model_status() -> dict[str, str]:
    """현재 모델 가용 상태 반환. /api/version에서 사용."""
    status = {"regex": "available"}

    ner_path = _get_ner_model_path()
    if not ner_path:
        status["ner"] = "not_configured"
    else:
        try:
            from transformers import pipeline  # noqa: F401
            status["ner"] = "available" if Path(ner_path).exists() else "path_not_found"
        except ImportError:
            status["ner"] = "not_installed"

    ai_path = _get_ai_model_path()
    if not ai_path:
        status["ai"] = "not_configured"
    else:
        try:
            import tensorflow as tf  # noqa: F401
            status["ai"] = "available" if Path(ai_path).exists() else "path_not_found"
        except ImportError:
            status["ai"] = "not_installed"

    return status


def _load_regex_func():
    from regex_detector import detect_patterns
    return detect_patterns


def _load_ner_func():
    """NER 함수 로드. 환경 변수 NER_MODEL_PATH 필요."""
    ner_path = _get_ner_model_path()
    if not ner_path:
        return None
    try:
        from transformers import pipeline
        pipe = pipeline(
            "ner",
            model=ner_path,
            tokenizer=ner_path,
            aggregation_strategy="simple",
        )
        return lambda text: pipe(text)
    except Exception:
        return None


def _load_ai_func():
    """AI 분류 함수 로드. 환경 변수 AI_MODEL_PATH 필요."""
    ai_path = _get_ai_model_path()
    if not ai_path or not Path(ai_path).exists():
        return None
    try:
        import tensorflow as tf
        model = tf.keras.models.load_model(ai_path)

        def predict(text: str):
            import tensorflow as tf
            preds = model.predict(tf.constant([text]), verbose=0)[0].tolist()
            labels = ["C", "S", "O"]
            prob_map = dict(zip(labels, preds))
            best = max(range(len(preds)), key=lambda i: preds[i])
            return labels[best], float(preds[best]), prob_map

        return predict
    except Exception:
        return None


# ── xlsx 처리 ──────────────────────────────────────────────────

def _run_xlsx(
    file_path: Path,
    original_name: str,
    deletion_mode: str,
    use_ner: bool,
    use_ai: bool,
) -> dict[str, Any]:
    """xlsx → applied 모드 처리 + downloadToken 생성."""
    import openpyxl
    from deidentify_target_builder import build_deidentify_plan
    from xlsx_deidentify_apply import apply_plan_to_xlsx

    regex_func = _load_regex_func()
    ner_func = _load_ner_func() if use_ner else None
    ai_func = _load_ai_func() if use_ai else None

    wb = openpyxl.load_workbook(str(file_path))
    detections: list[dict] = []
    order = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                text = cell.value
                if not isinstance(text, str) or not text.strip():
                    continue

                meta = {
                    "fileType": "xlsx",
                    "sheetName": sheet_name,
                    "cellRef": cell.coordinate,
                }

                def _g(obj, *names, default=None):
                    for n in names:
                        if isinstance(obj, dict) and n in obj:
                            return obj[n]
                        if hasattr(obj, n):
                            return getattr(obj, n)
                    return default

                for raw in (regex_func(text) or []):
                    detections.append({
                        "label": _g(raw, "label", default=""),
                        "matched": _g(raw, "value", "matched", default=""),
                        "grade": _g(raw, "grade", default="S"),
                        "action": _g(raw, "action", default="마스킹"),
                        "source": "regex",
                        "context": text,
                        "locationLabel": f"{sheet_name} {cell.coordinate}",
                        "locationMeta": meta,
                        "start": _g(raw, "start"),
                        "end": _g(raw, "end"),
                        "sensitiveType": _g(raw, "sensitive_type", "sensitiveType"),
                        "sensitiveCategory": _g(
                            raw, "sensitive_category", "sensitiveCategory",
                            default=_g(raw, "label", default=""),
                        ),
                        "reason": f"정규식 탐지: {_g(raw, 'label', default='')}",
                        "_order": order,
                    })
                    order += 1

                if ner_func:
                    for raw in (ner_func(text) or []):
                        entity = (
                            raw.get("entity_group") or raw.get("entity") or ""
                        ).upper().replace("B-", "").replace("I-", "")
                        if entity not in {"PERSON", "PER", "PS", "인명"}:
                            continue
                        if float(raw.get("score", 0)) < _get_ner_threshold():
                            continue
                        s, e = int(raw.get("start", 0)), int(raw.get("end", 0))
                        detections.append({
                            "label": "성명",
                            "matched": text[s:e],
                            "grade": "S",
                            "action": "마스킹",
                            "source": "ner",
                            "context": text,
                            "locationLabel": f"{sheet_name} {cell.coordinate}",
                            "locationMeta": meta,
                            "start": s, "end": e,
                            "sensitiveType": "개인정보",
                            "sensitiveCategory": "성명",
                            "reason": "NER 탐지: PERSON",
                            "_order": order,
                        })
                        order += 1

                if ai_func:
                    import logging as _log
                    try:
                        grade, confidence, prob_map = ai_func(text)
                        ai_threshold = _get_ai_threshold()
                        if grade != "O" and confidence >= ai_threshold:
                            prob_text = " / ".join(
                                f"{k}={v:.4f}" for k, v in prob_map.items()
                            )
                            detections.append({
                                "label": "민감정보",
                                "matched": "",
                                "grade": grade,
                                "action": "검토 필요",
                                "source": "ai",
                                "context": text,
                                "locationLabel": f"{sheet_name} {cell.coordinate}",
                                "locationMeta": meta,
                                "start": None, "end": None,
                                "sensitiveType": "문맥 기반 민감정보",
                                "sensitiveCategory": f"AI_{grade}",
                                "reason": (
                                    f"AI 문장분류 grade={grade} / "
                                    f"confidence={confidence:.4f} / "
                                    f"threshold={ai_threshold:.2f} / {prob_text}"
                                ),
                                "_order": order,
                            })
                            order += 1
                    except Exception as exc:
                        _log.getLogger(__name__).warning(
                            "[AI] %s %s 예측 실패: %s",
                            sheet_name, cell.coordinate, exc,
                        )

    wb.close()

    plan = build_deidentify_plan(detections)
    output_path = get_tmp_dir() / f"result_{secrets.token_hex(8)}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    result = apply_plan_to_xlsx(
        str(file_path), plan, str(output_path), deletion_mode,
    )

    d = result.to_dict()
    actual_output = Path(result.outputFilePath) if result.outputFilePath else output_path

    if actual_output.exists():
        d["downloadToken"] = create_download_token(actual_output, original_name)
    else:
        d["downloadToken"] = None

    d["outputFilePath"] = None
    return d


# ── guide 처리 ─────────────────────────────────────────────────

def _run_guide(
    file_path: Path,
    file_type: str,
    deletion_mode: str,
    use_ner: bool,
    use_ai: bool,
) -> dict[str, Any]:
    """docx/pptx/hwpx/pdf → guide 모드 처리."""
    regex_func = _load_regex_func()
    ner_func = _load_ner_func() if use_ner else None
    ai_func = _load_ai_func() if use_ai else None

    kwargs = dict(
        regex_detect_func=regex_func,
        ner_detect_func=ner_func,
        ai_predict_func=ai_func,
        deletion_mode=deletion_mode,
        ner_threshold=_get_ner_threshold(),
        ai_threshold=_get_ai_threshold(),
    )

    if file_type == "docx":
        from docx_detector import detect_and_build_guide_for_docx
        result = detect_and_build_guide_for_docx(str(file_path), **kwargs)
    elif file_type == "pptx":
        from pptx_detector import detect_and_build_guide_for_pptx
        result = detect_and_build_guide_for_pptx(str(file_path), **kwargs)
    elif file_type == "hwpx":
        from hwpx_detector import detect_and_build_guide_for_hwpx
        result = detect_and_build_guide_for_hwpx(str(file_path), **kwargs)
    elif file_type == "pdf":
        from pdf_detector import detect_and_build_guide_for_pdf
        result = detect_and_build_guide_for_pdf(str(file_path), **kwargs)
    else:
        raise ValueError(f"알 수 없는 fileType: {file_type}")

    d = result.to_dict()
    d["downloadToken"] = None
    return d


# ── 엔드포인트 ────────────────────────────────────────────────

@router.post(
    "/api/detect",
    summary="파일 비식별화 (통합)",
    description="""
업로드된 파일의 확장자를 자동으로 감지해 적절한 detector로 처리합니다.

**지원 형식:** xlsx, docx, pptx, hwpx, pdf

**applyMode별 응답 차이:**
- `guide` (docx/pptx/hwpx/pdf): `downloadToken`이 null
- `applied` (xlsx): `downloadToken`이 설정됨, `/api/download/{token}`으로 다운로드

**모델 경로 설정 (환경 변수):**
```
NER_MODEL_PATH=models/ner/KoELECTRA-small-v3-modu-ner
AI_MODEL_PATH=models/privacy_cso_char_keras_model.keras
```

**주의:**
- `useNer=true`인데 `NER_MODEL_PATH` 미설정 시 HTTP 503 반환
- `downloadToken`은 1회용, 1시간 후 만료
- `downloadToken` 저장소는 프로세스 메모리 기반 (단일 워커에서만 보장)
    """,
    tags=["detect"],
)
async def detect(
    file: UploadFile = File(...),
    deletionMode: str = Form(DEFAULT_DELETION_MODE),
    useNer: bool = Form(DEFAULT_USE_NER),
    useAi: bool = Form(DEFAULT_USE_AI),
    userId: str = Form(None),
) -> JSONResponse:

    filename = file.filename or "unknown"
    file_type = detect_file_type(filename)

    if file_type is None:
        return _error(
            400, "UNSUPPORTED_FILE_TYPE",
            "지원하지 않는 파일 형식입니다. 지원 형식: xlsx, docx, pptx, hwpx, pdf",
            filename=filename,
        )

    # 모델 미설정 검사
    if useNer and not _get_ner_model_path():
        return _error(
            503, "NER_MODEL_NOT_CONFIGURED",
            "NER 모델 경로가 설정되지 않았습니다. "
            "환경 변수 NER_MODEL_PATH를 설정하거나 useNer=false로 요청하세요.",
        )
    if useAi and not _get_ai_model_path():
        return _error(
            503, "AI_MODEL_NOT_CONFIGURED",
            "AI 모델 경로가 설정되지 않았습니다. "
            "환경 변수 AI_MODEL_PATH를 설정하거나 useAi=false로 요청하세요.",
        )

    content = await file.read()
    filesize = len(content)
    suffix = Path(filename).suffix.lower()
    tmp_path = save_upload_tmp(content, suffix)

    try:
        if file_type == "xlsx":
            result_dict = _run_xlsx(
                tmp_path, original_name=filename,
                deletion_mode=deletionMode, use_ner=useNer, use_ai=useAi,
            )
        else:
            result_dict = _run_guide(
                tmp_path, file_type=file_type,
                deletion_mode=deletionMode, use_ner=useNer, use_ai=useAi,
            )
        return _success(result_dict, filename=filename, filesize=filesize)

    except Exception as exc:
        return _error(
            500, "INTERNAL_ERROR",
            f"처리 중 오류가 발생했습니다: {type(exc).__name__}: {exc}",
            fileType=file_type,
        )
    finally:
        cleanup_upload(tmp_path)


# ── 다운로드 엔드포인트 ────────────────────────────────────────

@router.get(
    "/api/download/{token}",
    summary="xlsx 비식별화 파일 다운로드",
    description="""
`/api/detect` 응답의 `downloadToken`으로 비식별화된 xlsx 파일을 다운로드합니다.

- 토큰 유효 시간: **1시간**
- **1회용**: 다운로드 후 토큰 즉시 만료
- 만료 또는 없는 토큰: HTTP 404
    """,
    tags=["detect"],
)
async def download(token: str) -> FileResponse:

    entry = resolve_download_token(token)
    if entry is None:
        raise HTTPException(status_code=404, detail="다운로드 토큰이 없거나 만료되었습니다.")
    if not entry.file_path.exists():
        revoke_token(token)
        raise HTTPException(status_code=404, detail="파일이 존재하지 않습니다.")

    file_path = entry.file_path
    original_name = entry.original_name
    _token_store_remove(token)

    stem = Path(original_name).stem
    download_name = f"{stem}_deidentified.xlsx"

    def _cleanup():
        try:
            file_path.unlink(missing_ok=True)
        except Exception:
            pass

    return FileResponse(
        path=str(file_path),
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(_cleanup),
    )


# ── cleanup 헬퍼 (필요 시 서버 시작 시 호출) ──────────────────

def startup_cleanup() -> None:
    """서버 시작 시 만료 토큰 + 고아 파일 정리."""
    expired = cleanup_expired_tokens()
    orphans = cleanup_orphan_files()
    if expired or orphans:
        import sys
        print(
            f"[startup_cleanup] 만료 토큰 {expired}건, "
            f"고아 파일 {orphans}건 정리 완료",
            file=sys.stderr,
        )


# cleanup_upload는 files.py에서 import
from api.files import cleanup_upload  # noqa: E402