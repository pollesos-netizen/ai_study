"""
xlsx 파일 단위 비식별화 Apply

목적:
- DeidentifyPlan의 xlsx auto_targets를 실제 xlsx 파일 셀 값에 적용합니다.
- 결과는 파일 형식 공통 구조인 CommonApplyResult로 반환합니다.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any
import unicodedata

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

try:
    from src.common_apply_result import (
        CommonApplyItem,
        CommonApplyResult,
        build_summary,
        make_review_items,
    )
    from src.deidentify_apply import apply_targets_to_text
    from src.deidentify_target_builder import DeidentifyPlan, DeidentifyTarget
except ModuleNotFoundError:
    from common_apply_result import (
        CommonApplyItem,
        CommonApplyResult,
        build_summary,
        make_review_items,
    )
    from deidentify_apply import apply_targets_to_text
    from deidentify_target_builder import DeidentifyPlan, DeidentifyTarget


def normalize_nfc(value: Any) -> str:
    return unicodedata.normalize("NFC", str(value))


def make_output_path(input_path: str | Path, output_path: str | Path | None = None) -> str:
    """
    output_path가 없으면 original_deidentified.xlsx 규칙으로 생성합니다.
    기존 파일이 있으면 _1, _2 suffix를 붙입니다.
    """
    if output_path is not None:
        return str(output_path)

    input_path = Path(input_path)
    base = input_path.with_name(f"{input_path.stem}_deidentified{input_path.suffix}")

    if not base.exists():
        return str(base)

    index = 1
    while True:
        candidate = input_path.with_name(
            f"{input_path.stem}_deidentified_{index}{input_path.suffix}"
        )
        if not candidate.exists():
            return str(candidate)
        index += 1


def target_file_type(target: DeidentifyTarget) -> str:
    return str((target.location_meta or {}).get("fileType") or "").lower()


def filter_xlsx_targets(plan: DeidentifyPlan) -> list[DeidentifyTarget]:
    """
    fileType이 xlsx인 auto target만 선별합니다.
    """
    return [
        target
        for target in plan.auto_targets
        if target_file_type(target) == "xlsx"
    ]


def normalize_sheet_name(name: Any) -> str:
    return unicodedata.normalize("NFC", str(name))


def find_worksheet(workbook, sheet_name: str):
    """
    NFC 정규화된 sheetName으로 worksheet를 찾습니다.
    실제 접근은 workbook에 존재하는 원본 sheetName으로 수행합니다.
    """
    normalized_target = normalize_sheet_name(sheet_name)

    for actual_name in workbook.sheetnames:
        if normalize_sheet_name(actual_name) == normalized_target:
            return workbook[actual_name]

    return None


def is_formula_cell(cell: Cell) -> bool:
    return cell.data_type == "f"


def is_string_cell(cell: Cell) -> bool:
    return isinstance(cell.value, str) and not is_formula_cell(cell)


def cell_type_name(cell: Cell) -> str:
    if cell.value is None:
        return "empty"
    if is_formula_cell(cell):
        return "formula"
    if isinstance(cell.value, bool):
        return "boolean"
    if isinstance(cell.value, (int, float)):
        return "number"
    if isinstance(cell.value, str):
        return "string"
    return type(cell.value).__name__


def get_location_label(target: DeidentifyTarget, fallback: str) -> str:
    return target.location_label or fallback


def get_sheet_and_cell(target: DeidentifyTarget) -> tuple[str | None, str | None]:
    meta = target.location_meta or {}
    return meta.get("sheetName"), meta.get("cellRef")


def merged_cell_status(ws, cell_ref: str) -> tuple[bool, bool, str | None, str | None]:
    """
    병합 셀 여부를 확인합니다.

    Returns:
        is_merged:
            병합 범위에 포함되는지
        is_top_left:
            병합 범위의 좌상단 셀인지
        merged_range:
            병합 범위 문자열
        top_left:
            좌상단 cellRef
    """
    row, col = coordinate_to_tuple(cell_ref)

    for merged_range in ws.merged_cells.ranges:
        in_range = (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        )

        if not in_range:
            continue

        top_left = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
        return True, cell_ref == top_left, str(merged_range), top_left

    return False, False, None, None

def group_targets_by_sheet_cell(
    targets: list[DeidentifyTarget],
) -> tuple[dict[tuple[str, str], list[DeidentifyTarget]], list[CommonApplyItem], list[str]]:
    """
    xlsx target을 (sheetName, cellRef) 기준으로 묶습니다.

    sheetName/cellRef가 없으면 skipped CommonApplyItem으로 변환합니다.
    """
    grouped: dict[tuple[str, str], list[DeidentifyTarget]] = {}
    skipped_items: list[CommonApplyItem] = []
    warnings: list[str] = []

    for target in targets:
        sheet_name, cell_ref = get_sheet_and_cell(target)
        label = target.label or ""

        if not sheet_name:
            warning = f"{target.location_label}: sheetName이 없어 자동 적용을 건너뛰었습니다."
            warnings.append(warning)
            skipped_items.append(
                CommonApplyItem(
                    locationLabel=target.location_label,
                    locationMeta=target.location_meta or {},
                    label=label,
                    action=target.action,
                    originalText=target.context or "",
                    appliedText=target.context or "",
                    status="skipped",
                    appliedTargetCount=0,
                    skippedTargetCount=1,
                    warnings=[warning],
                )
            )
            continue

        if not cell_ref:
            warning = f"{target.location_label}: cellRef가 없어 자동 적용을 건너뛰었습니다."
            warnings.append(warning)
            skipped_items.append(
                CommonApplyItem(
                    locationLabel=target.location_label,
                    locationMeta=target.location_meta or {},
                    label=label,
                    action=target.action,
                    originalText=target.context or "",
                    appliedText=target.context or "",
                    status="skipped",
                    appliedTargetCount=0,
                    skippedTargetCount=1,
                    warnings=[warning],
                )
            )
            continue

        grouped.setdefault((str(sheet_name), str(cell_ref)), []).append(target)

    return grouped, skipped_items, warnings


def labels_for_targets(targets: list[DeidentifyTarget]) -> str:
    labels: list[str] = []

    for target in targets:
        if target.label and target.label not in labels:
            labels.append(target.label)

    return ", ".join(labels)


def actions_for_targets(targets: list[DeidentifyTarget]) -> str:
    actions: list[str] = []

    for target in targets:
        if target.action and target.action not in actions:
            actions.append(target.action)

    return ", ".join(actions)


def validate_slice_against_matched(
    cell_value: str,
    target: DeidentifyTarget,
) -> str | None:
    """
    실제 cell_value[start:end]와 matched가 직접 일치하는지 검증합니다.

    NFC 정규화 후에만 일치하는 경우도 자동 적용하지 않고 warning으로 처리합니다.
    """
    if target.start is None or target.end is None:
        return "start 또는 end가 None입니다."

    if target.start < 0 or target.end > len(cell_value) or target.start >= target.end:
        return (
            "start/end 범위가 현재 셀 값에 유효하지 않습니다: "
            f"start={target.start}, end={target.end}, cell_len={len(cell_value)}"
        )

    actual = cell_value[target.start:target.end]
    matched = target.matched or ""

    if actual == matched:
        return None

    if normalize_nfc(actual) == normalize_nfc(matched):
        return (
            "cell_value[start:end]와 matched가 원문 기준으로 일치하지 않습니다. "
            "NFC 정규화 후에만 일치하므로 인덱스 기준 불일치 가능성이 있어 자동 적용을 건너뜁니다: "
            f"actual={actual!r}, matched={matched!r}, start={target.start}, end={target.end}"
        )

    return (
        "cell_value[start:end]와 matched가 일치하지 않아 자동 적용을 건너뜁니다: "
        f"actual={actual!r}, matched={matched!r}, start={target.start}, end={target.end}"
    )


def make_status(applied_count: int, skipped_count: int) -> str:
    if applied_count > 0 and skipped_count == 0:
        return "applied"
    if applied_count > 0 and skipped_count > 0:
        return "partial"
    return "skipped"


def apply_targets_to_cell(
    ws,
    sheet_name: str,
    cell_ref: str,
    targets: list[DeidentifyTarget],
    *,
    deletion_mode: str,
) -> CommonApplyItem:
    """
    특정 셀에 속한 target 목록을 셀 값에 적용합니다.
    """
    representative = targets[0]
    location_label = representative.location_label or f"{sheet_name} 탭 {cell_ref} 셀"
    location_meta = representative.location_meta or {
        "fileType": "xlsx",
        "sheetName": sheet_name,
        "cellRef": cell_ref,
    }

    warnings: list[str] = []

    cell = ws[cell_ref]

    if isinstance(cell, MergedCell):
        warning = (
            f"{location_label}: 병합 셀의 좌상단 셀이 아니므로 자동 적용을 건너뛰었습니다. "
            f"cellRef={cell_ref}"
        )
        warnings.append(warning)
        return CommonApplyItem(
            locationLabel=location_label,
            locationMeta=location_meta,
            label=labels_for_targets(targets),
            action=actions_for_targets(targets),
            originalText="",
            appliedText="",
            status="skipped",
            appliedTargetCount=0,
            skippedTargetCount=len(targets),
            warnings=warnings,
        )

    is_merged, is_top_left, merged_range, top_left = merged_cell_status(ws, cell_ref)
    if is_merged and not is_top_left:
        warning = (
            f"{location_label}: 병합 셀의 좌상단 셀이 아니므로 자동 적용을 건너뛰었습니다. "
            f"mergedRange={merged_range}, topLeft={top_left}"
        )
        warnings.append(warning)
        return CommonApplyItem(
            locationLabel=location_label,
            locationMeta=location_meta,
            label=labels_for_targets(targets),
            action=actions_for_targets(targets),
            originalText=str(cell.value or ""),
            appliedText=str(cell.value or ""),
            status="skipped",
            appliedTargetCount=0,
            skippedTargetCount=len(targets),
            warnings=warnings,
        )

    if is_formula_cell(cell):
        warning = (
            f"{location_label}: 수식 셀에 detection이 있어 자동 적용을 건너뛰었습니다. "
            f"cellType=formula, formula={cell.value}"
        )
        warnings.append(warning)
        return CommonApplyItem(
            locationLabel=location_label,
            locationMeta=location_meta,
            label=labels_for_targets(targets),
            action=actions_for_targets(targets),
            originalText=str(cell.value or ""),
            appliedText=str(cell.value or ""),
            status="skipped",
            appliedTargetCount=0,
            skippedTargetCount=len(targets),
            warnings=warnings,
        )

    if not is_string_cell(cell):
        warning = (
            f"{location_label}: 비문자열 셀에 detection이 있어 자동 적용을 건너뛰었습니다. "
            f"cellType={cell_type_name(cell)}, cellValue={cell.value}"
        )
        warnings.append(warning)
        return CommonApplyItem(
            locationLabel=location_label,
            locationMeta=location_meta,
            label=labels_for_targets(targets),
            action=actions_for_targets(targets),
            originalText=str(cell.value or ""),
            appliedText=str(cell.value or ""),
            status="skipped",
            appliedTargetCount=0,
            skippedTargetCount=len(targets),
            warnings=warnings,
        )

    cell_value = cell.value

    if any(target.context is not None and normalize_nfc(target.context) != normalize_nfc(cell_value) for target in targets):
        warnings.append(
            f"{location_label}: target.context와 실제 셀 값이 다릅니다. "
            "cell_value 기준으로 slice 검증 후 적용 여부를 판단합니다."
        )

    valid_targets: list[DeidentifyTarget] = []
    skipped_count = 0

    for target in targets:
        slice_error = validate_slice_against_matched(cell_value, target)
        if slice_error is not None:
            warnings.append(f"{location_label}: {slice_error}")
            skipped_count += 1
            continue

        valid_targets.append(target)

    if valid_targets:
        apply_result = apply_targets_to_text(
            cell_value,
            valid_targets,
            deletion_mode=deletion_mode,
        )
        cell.value = apply_result.applied_text
        warnings.extend(apply_result.warnings)
        applied_count = len(apply_result.applied_targets)
        skipped_count += len(apply_result.skipped_targets)
        applied_text = apply_result.applied_text
    else:
        applied_count = 0
        applied_text = cell_value

    return CommonApplyItem(
        locationLabel=location_label,
        locationMeta=location_meta,
        label=labels_for_targets(targets),
        action=actions_for_targets(targets),
        originalText=cell_value,
        appliedText=applied_text,
        status=make_status(applied_count, skipped_count),
        appliedTargetCount=applied_count,
        skippedTargetCount=skipped_count,
        warnings=warnings,
    )


def apply_plan_to_xlsx(
    input_path: str,
    plan: DeidentifyPlan,
    output_path: str | None = None,
    deletion_mode: str = "delete",
) -> CommonApplyResult:
    """
    DeidentifyPlan을 xlsx 파일에 적용합니다.

    모든 파일 형식별 Apply 함수는 이 공통 시그니처 패턴을 따릅니다.
    """
    resolved_output_path = make_output_path(input_path, output_path)

    wb = load_workbook(input_path, data_only=False)

    xlsx_targets = filter_xlsx_targets(plan)
    grouped, skipped_items, global_warnings = group_targets_by_sheet_cell(xlsx_targets)

    auto_results: list[CommonApplyItem] = []
    auto_results.extend(skipped_items)

    for (sheet_name, cell_ref), targets in grouped.items():
        ws = find_worksheet(wb, sheet_name)

        if ws is None:
            warning = f"{sheet_name} 시트를 찾을 수 없어 자동 적용을 건너뛰었습니다."
            global_warnings.append(warning)
            auto_results.append(
                CommonApplyItem(
                    locationLabel=targets[0].location_label,
                    locationMeta=targets[0].location_meta or {},
                    label=labels_for_targets(targets),
                    action=actions_for_targets(targets),
                    originalText=targets[0].context or "",
                    appliedText=targets[0].context or "",
                    status="skipped",
                    appliedTargetCount=0,
                    skippedTargetCount=len(targets),
                    warnings=[warning],
                )
            )
            continue

        item = apply_targets_to_cell(
            ws,
            sheet_name,
            cell_ref,
            targets,
            deletion_mode=deletion_mode,
        )
        auto_results.append(item)

    wb.save(resolved_output_path)

    review_items = make_review_items(plan.review_targets)
    summary = build_summary(auto_results, review_items, global_warnings)

    return CommonApplyResult(
        fileType="xlsx",
        inputFilePath=str(input_path),
        outputFilePath=resolved_output_path,
        autoResults=auto_results,
        reviewTargets=review_items,
        warnings=global_warnings,
        summary=summary,
    )
