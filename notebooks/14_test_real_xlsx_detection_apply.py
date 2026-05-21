"""
실제 xlsx 파일 대상 regex + NER + AI 탐지 후 Apply 통합 테스트

실행 예시:
    python notebooks/14_test_real_xlsx_detection_apply.py "C:/path/to/input.xlsx" \
        --ner-model-path "models/hf/KoELECTRA-small-v3-modu-ner" \
        --ai-model-path "models/privacy_cso_char_keras_model.keras"

설명:
- 실제 xlsx 파일의 모든 문자열 셀을 순회합니다.
- regex 탐지는 src/regex_detector.py의 공통 정규식 탐지기를 사용합니다.
- NER 모델 경로가 제공되면 성명 탐지를 수행합니다.
- AI 모델 경로가 제공되면 문장/셀 단위 검토 필요 여부를 판단합니다.
- auto_targets는 xlsx 파일에 적용됩니다.
- AI review_targets는 자동 적용하지 않고 CommonApplyResult.reviewTargets에 보존됩니다.

주의:
- 이 스크립트는 모델 성능 평가가 아니라 파이프라인 연결 검증용입니다.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import argparse
import inspect
import sys
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SRC_DIR = PROJECT_ROOT / "src"

if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from deidentify_target_builder import DeidentifyPlan, DeidentifyTarget
from xlsx_deidentify_apply import apply_plan_to_xlsx

try:
    import regex_detector
except ImportError as exc:
    raise ImportError(
        "src/regex_detector.py를 import할 수 없습니다. "
        "공통 정규식 탐지기를 사용하도록 프로젝트 경로와 파일명을 확인하세요."
    ) from exc


# -----------------------------
# 기본 설정
# -----------------------------

DEFAULT_NER_THRESHOLD = 0.8
DEFAULT_AI_THRESHOLD = 0.6


@dataclass
class ParsedCell:
    sheet_name: str
    cell_ref: str
    text: str

    @property
    def location_label(self) -> str:
        return f"{self.sheet_name} 탭 {self.cell_ref} 셀"

    @property
    def location_meta(self) -> dict[str, Any]:
        return {
            "fileType": "xlsx",
            "sheetName": self.sheet_name,
            "cellRef": self.cell_ref,
        }


# -----------------------------
# 공통 target 생성
# -----------------------------

def make_target(
    *,
    label: str,
    matched: str,
    start: int | None,
    end: int | None,
    source: str,
    grade: str,
    action: str,
    cell: ParsedCell,
    sensitive_type: str,
    sensitive_category: str,
    reason: str,
) -> DeidentifyTarget:
    return DeidentifyTarget(
        label=label,
        matched=matched,
        action=action,
        location_label=cell.location_label,
        location_meta=cell.location_meta,
        start=start,
        end=end,
        source=source,
        reason=reason,
        grade=grade,
        sensitive_type=sensitive_type,
        sensitive_category=sensitive_category,
        context=cell.text,
        order=0,
    )


# -----------------------------
# regex_detector.py 어댑터
# -----------------------------

def _get_attr_or_key(obj: Any, *names: str, default: Any = None) -> Any:
    """
    dict / dataclass / 일반 객체에서 여러 후보 이름 중 첫 값을 가져옵니다.
    """
    for name in names:
        if isinstance(obj, dict) and name in obj:
            return obj.get(name)

        if hasattr(obj, name):
            return getattr(obj, name)

    return default


def _find_regex_detect_function():
    """
    regex_detector.py의 함수명을 프로젝트 구현에 맞게 자동 탐색합니다.

    우선순위:
    1. detect_regex
    2. detect_regex_patterns
    3. detect_patterns
    4. detect
    5. find_detections
    6. find_matches
    """
    candidates = [
        "detect_regex",
        "detect_regex_patterns",
        "detect_patterns",
        "detect",
        "find_detections",
        "find_matches",
    ]

    for name in candidates:
        func = getattr(regex_detector, name, None)
        if callable(func):
            return func

    available = [
        name
        for name, value in vars(regex_detector).items()
        if callable(value) and not name.startswith("_")
    ]

    raise AttributeError(
        "regex_detector.py에서 사용할 수 있는 탐지 함수를 찾지 못했습니다. "
        f"후보={candidates}, 사용 가능 함수={available}"
    )


def _call_regex_detector(func, cell: ParsedCell):
    """
    regex_detector 함수 시그니처가 조금 달라도 호출 가능하도록 방어적으로 호출합니다.
    """
    sig = inspect.signature(func)
    params = sig.parameters

    kwargs: dict[str, Any] = {}

    if "text" in params:
        kwargs["text"] = cell.text

    if "location_meta" in params:
        kwargs["location_meta"] = cell.location_meta

    if "locationMeta" in params:
        kwargs["locationMeta"] = cell.location_meta

    if "location_label" in params:
        kwargs["location_label"] = cell.location_label

    if "locationLabel" in params:
        kwargs["locationLabel"] = cell.location_label

    if "context" in params:
        kwargs["context"] = cell.text

    if kwargs:
        try:
            return func(**kwargs)
        except TypeError:
            pass

    # 가장 흔한 단일 인자 형태
    try:
        return func(cell.text)
    except TypeError:
        pass

    # 일부 구현이 dict 입력을 받을 가능성
    return func(
        {
            "text": cell.text,
            "context": cell.text,
            "locationMeta": cell.location_meta,
            "locationLabel": cell.location_label,
        }
    )


def _normalize_regex_detection_to_target(raw: Any, cell: ParsedCell) -> DeidentifyTarget | None:
    """
    regex_detector.py의 반환 형식이 dict/dataclass 등이어도 DeidentifyTarget으로 변환합니다.
    """
    label = _get_attr_or_key(raw, "label", "name", "detectedLabel", default=None)
    matched = _get_attr_or_key(raw, "matched", "match", "value", "text", "detectedText", default=None)

    start = _get_attr_or_key(raw, "start", "startIndex", "begin", default=None)
    end = _get_attr_or_key(raw, "end", "endIndex", "finish", default=None)

    grade = _get_attr_or_key(raw, "grade", "csoGrade", "cso_grade", default="S")
    action = _get_attr_or_key(raw, "action", "recommendedAction", default="마스킹")

    reason = _get_attr_or_key(raw, "reason", "desc", "description", default=None)

    sensitive_type = _get_attr_or_key(
        raw,
        "sensitive_type",
        "sensitiveType",
        "type",
        default="정규식 기반 탐지",
    )
    sensitive_category = _get_attr_or_key(
        raw,
        "sensitive_category",
        "sensitiveCategory",
        "category",
        default=label or "정규식",
    )

    # 일부 구현은 span을 튜플로 줄 수 있음
    span = _get_attr_or_key(raw, "span", default=None)
    if span and (start is None or end is None):
        try:
            start, end = span
        except Exception:
            pass

    if matched is None and start is not None and end is not None:
        try:
            matched = cell.text[int(start):int(end)]
        except Exception:
            matched = ""

    if label is None or matched is None or start is None or end is None:
        return None

    try:
        start_int = int(start)
        end_int = int(end)
    except Exception:
        return None

    return make_target(
        label=str(label),
        matched=str(matched),
        start=start_int,
        end=end_int,
        source="regex",
        grade=str(grade),
        action=str(action),
        cell=cell,
        sensitive_type=str(sensitive_type),
        sensitive_category=str(sensitive_category),
        reason=str(reason or f"공통 regex_detector 탐지: {label}"),
    )


_REGEX_DETECT_FUNC = None


def detect_regex_targets(cell: ParsedCell) -> list[DeidentifyTarget]:
    """
    src/regex_detector.py의 공통 탐지기를 사용해 regex target을 생성합니다.
    """
    global _REGEX_DETECT_FUNC

    if _REGEX_DETECT_FUNC is None:
        _REGEX_DETECT_FUNC = _find_regex_detect_function()

    raw_results = _call_regex_detector(_REGEX_DETECT_FUNC, cell)

    if raw_results is None:
        return []

    # 단일 dict/dataclass 반환 방어
    if isinstance(raw_results, dict):
        raw_iterable = [raw_results]
    else:
        try:
            raw_iterable = list(raw_results)
        except TypeError:
            raw_iterable = [raw_results]

    targets: list[DeidentifyTarget] = []

    for raw in raw_iterable:
        target = _normalize_regex_detection_to_target(raw, cell)
        if target is not None:
            targets.append(target)

    return targets


# -----------------------------
# NER 탐지
# -----------------------------

def load_ner_pipeline(model_path: str | None):
    if not model_path:
        return None

    try:
        from transformers import pipeline
    except ImportError:
        print("[NER] transformers가 설치되어 있지 않아 NER 탐지를 건너뜁니다.")
        return None

    print(f"[NER] 모델 로드: {model_path}")
    return pipeline(
        "ner",
        model=model_path,
        tokenizer=model_path,
        aggregation_strategy="simple",
    )


def normalize_ner_label(label: str | None) -> str | None:
    if not label:
        return None

    label = label.replace("B-", "").replace("I-", "").upper()

    if label in {"PS", "PER", "PERSON", "인명"}:
        return "PERSON"

    return None


def detect_ner(
    cell: ParsedCell,
    ner_pipe,
    *,
    threshold: float,
) -> list[DeidentifyTarget]:
    if ner_pipe is None:
        return []

    try:
        raw_outputs = ner_pipe(cell.text)
    except Exception as exc:
        print(f"[NER] 탐지 실패: {cell.location_label} / {exc}")
        return []

    targets: list[DeidentifyTarget] = []

    for raw in raw_outputs:
        original_label = raw.get("entity_group") or raw.get("entity")
        normalized = normalize_ner_label(original_label)

        if normalized != "PERSON":
            continue

        score = float(raw.get("score") or 0.0)
        if score < threshold:
            continue

        start = raw.get("start")
        end = raw.get("end")
        word = raw.get("word") or ""

        if start is None or end is None:
            continue

        actual = cell.text[int(start):int(end)]

        targets.append(
            make_target(
                label="성명",
                matched=actual or word,
                start=int(start),
                end=int(end),
                source="ner",
                grade="S",
                action="마스킹",
                cell=cell,
                sensitive_type="개인정보",
                sensitive_category="성명",
                reason=(
                    "NER 모델이 PERSON 개체로 탐지"
                    f" / original_label={original_label}"
                    f" / confidence={score:.4f}"
                    f" / threshold={threshold:.2f}"
                ),
            )
        )

    return targets


# -----------------------------
# AI 문장/셀 분류
# -----------------------------

def load_ai_model(model_path: str | None):
    if not model_path:
        return None

    try:
        import tensorflow as tf
    except ImportError:
        print("[AI] tensorflow가 설치되어 있지 않아 AI 분류를 건너뜁니다.")
        return None

    print(f"[AI] 모델 로드: {model_path}")
    return tf.keras.models.load_model(model_path)


def predict_ai_grade(ai_model, text: str) -> tuple[str | None, float | None, dict[str, float]]:
    """
    8주차 Keras char model을 가정합니다.

    모델 출력 순서는 C, S, O로 가정합니다.
    """
    if ai_model is None:
        return None, None, {}

    try:
        import tensorflow as tf

        inputs = tf.constant([text])
        predictions = ai_model.predict(inputs, verbose=0)
        probs = predictions[0].tolist()

        labels = ["C", "S", "O"]
        prob_map = {
            label: float(prob)
            for label, prob in zip(labels, probs)
        }

        best_index = max(range(len(probs)), key=lambda idx: probs[idx])
        best_label = labels[best_index]
        best_prob = float(probs[best_index])

        return best_label, best_prob, prob_map

    except Exception as exc:
        print(f"[AI] 예측 실패: {exc}")
        return None, None, {}


def detect_ai_review(
    cell: ParsedCell,
    ai_model,
    *,
    threshold: float,
) -> list[DeidentifyTarget]:
    if ai_model is None:
        return []

    grade, confidence, prob_map = predict_ai_grade(ai_model, cell.text)

    if grade is None or confidence is None:
        return []

    # O 또는 threshold 미만은 review target으로 만들지 않음
    if grade == "O" or confidence < threshold:
        return []

    prob_text = ", ".join(
        f"{label}={prob:.3f}"
        for label, prob in prob_map.items()
    )

    return [
        make_target(
            label="민감정보",
            matched="",
            start=None,
            end=None,
            source="ai",
            grade=grade,
            action="검토 필요",
            cell=cell,
            sensitive_type="문맥 기반 민감정보",
            sensitive_category=f"AI_{grade}",
            reason=(
                f"AI 문장분류 결과 grade={grade}"
                f" / confidence={confidence:.4f}"
                f" / threshold={threshold:.2f}"
                f" / probs=({prob_text})"
            ),
        )
    ]


# -----------------------------
# xlsx 순회 및 plan 생성
# -----------------------------

def iter_string_cells(input_path: str) -> list[ParsedCell]:
    wb = load_workbook(input_path, data_only=False)

    cells: list[ParsedCell] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cells.append(
                        ParsedCell(
                            sheet_name=ws.title,
                            cell_ref=cell.coordinate,
                            text=cell.value,
                        )
                    )

    return cells


def build_plan_from_xlsx(
    input_path: str,
    *,
    ner_model_path: str | None,
    ai_model_path: str | None,
    ner_threshold: float,
    ai_threshold: float,
    debug_cells: bool = False,
) -> tuple[DeidentifyPlan, dict[str, int]]:
    ner_pipe = load_ner_pipeline(ner_model_path)
    ai_model = load_ai_model(ai_model_path)

    cells = iter_string_cells(input_path)

    auto_targets: list[DeidentifyTarget] = []
    review_targets: list[DeidentifyTarget] = []

    regex_count = 0
    ner_count = 0
    ai_count = 0

    for order, cell in enumerate(cells):
        regex_targets = detect_regex_targets(cell)
        ner_targets = detect_ner(cell, ner_pipe, threshold=ner_threshold)
        ai_targets = detect_ai_review(cell, ai_model, threshold=ai_threshold)

        for target in [*regex_targets, *ner_targets, *ai_targets]:
            target.order = order

        auto_targets.extend(regex_targets)
        auto_targets.extend(ner_targets)
        review_targets.extend(ai_targets)

        regex_count += len(regex_targets)
        ner_count += len(ner_targets)
        ai_count += len(ai_targets)

        if debug_cells:
            print(f"[Cell] {cell.location_label}: {cell.text!r}")
            print(
                f"  regex={len(regex_targets)}, "
                f"ner={len(ner_targets)}, "
                f"ai={len(ai_targets)}"
            )
            for target in [*regex_targets, *ner_targets, *ai_targets]:
                print(
                    f"    - {target.source}/{target.label}: "
                    f"{target.matched or '문장 전체'} "
                    f"({target.start},{target.end})"
                )

    stats = {
        "stringCellCount": len(cells),
        "regexTargetCount": regex_count,
        "nerTargetCount": ner_count,
        "aiReviewTargetCount": ai_count,
        "autoTargetCount": len(auto_targets),
        "reviewTargetCount": len(review_targets),
    }

    return (
        DeidentifyPlan(
            auto_targets=auto_targets,
            review_targets=review_targets,
            summary_grade=None,
        ),
        stats,
    )


# -----------------------------
# 출력
# -----------------------------

def print_plan_stats(stats: dict[str, int]) -> None:
    print("\n=== 실제 xlsx 탐지 통계 ===")
    for key, value in stats.items():
        print(f"{key}: {value}")


def print_common_result(result) -> None:
    print("\n=== 실제 xlsx Detection + Apply 결과 ===")
    print(f"fileType: {result.fileType}")
    print(f"inputFilePath: {result.inputFilePath}")
    print(f"outputFilePath: {result.outputFilePath}")

    print("\nsummary:")
    print(f"  totalLocations: {result.summary.totalLocations}")
    print(f"  appliedLocations: {result.summary.appliedLocations}")
    print(f"  partialLocations: {result.summary.partialLocations}")
    print(f"  skippedLocations: {result.summary.skippedLocations}")
    print(f"  totalWarnings: {result.summary.totalWarnings}")
    print(f"  autoTargetCount: {result.summary.autoTargetCount}")
    print(f"  reviewTargetCount: {result.summary.reviewTargetCount}")

    print(f"\nautoResults: {len(result.autoResults)}")
    for item in result.autoResults[:50]:
        print(
            f"  - {item.locationLabel} / {item.label} / status={item.status} / "
            f"applied={item.appliedTargetCount}, skipped={item.skippedTargetCount}"
        )
        print(f"    original: {item.originalText}")
        print(f"    applied : {item.appliedText}")
        for warning in item.warnings:
            print(f"    warning: {warning}")

    if len(result.autoResults) > 50:
        print(f"  ... autoResults {len(result.autoResults) - 50}건 생략")

    print(f"\nreviewTargets: {len(result.reviewTargets)}")
    for review in result.reviewTargets[:50]:
        print(
            f"  - {review.locationLabel} / {review.label} / "
            f"action={review.action} / reason={review.reason}"
        )
        print(f"    context: {review.context}")

    if len(result.reviewTargets) > 50:
        print(f"  ... reviewTargets {len(result.reviewTargets) - 50}건 생략")

    print(f"\nglobal warnings: {len(result.warnings)}")
    for warning in result.warnings:
        print(f"  - {warning}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_path", help="실제 xlsx 파일 경로")
    parser.add_argument("--output-path", default=None, help="결과 xlsx 파일 경로")
    parser.add_argument(
        "--deletion-mode",
        default="mark",
        choices=["delete", "mark"],
        help="delete: 실제 삭제, mark: preview용 (삭제됨) 표시",
    )
    parser.add_argument(
        "--ner-model-path",
        default=None,
        help="Hugging Face NER 모델 로컬 경로. 생략하면 NER 탐지 skip",
    )
    parser.add_argument(
        "--ai-model-path",
        default=None,
        help="Keras AI 문장분류 모델 경로. 생략하면 AI 탐지 skip",
    )
    parser.add_argument(
        "--ner-threshold",
        type=float,
        default=DEFAULT_NER_THRESHOLD,
    )
    parser.add_argument(
        "--ai-threshold",
        type=float,
        default=DEFAULT_AI_THRESHOLD,
    )
    parser.add_argument(
        "--debug-cells",
        action="store_true",
        help="모든 문자열 셀의 regex/NER/AI 탐지 건수를 출력합니다.",
    )

    args = parser.parse_args()

    input_path = Path(args.input_path)

    if not input_path.exists():
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {input_path}")

    print("=== 실제 xlsx regex + NER + AI 탐지 후 Apply 통합 테스트 ===")
    print(f"input_path: {input_path}")
    print(f"ner_model_path: {args.ner_model_path}")
    print(f"ai_model_path: {args.ai_model_path}")
    print(f"ner_threshold: {args.ner_threshold}")
    print(f"ai_threshold: {args.ai_threshold}")
    print(f"deletion_mode: {args.deletion_mode}")
    if args.deletion_mode == "mark":
        print("삭제 대상은 사용자 확인을 위해 '(삭제됨)'으로 표시됩니다.")
    else:
        print("삭제 대상은 실제 삭제되어 빈 문자열로 제거됩니다.")
    plan, stats = build_plan_from_xlsx(
        str(input_path),
        ner_model_path=args.ner_model_path,
        ai_model_path=args.ai_model_path,
        ner_threshold=args.ner_threshold,
        ai_threshold=args.ai_threshold,
        debug_cells=args.debug_cells,
    )

    print_plan_stats(stats)

    result = apply_plan_to_xlsx(
        str(input_path),
        plan,
        output_path=args.output_path,
        deletion_mode=args.deletion_mode,
    )

    print_common_result(result)

    print("\n=== 통합 테스트 완료 ===")


if __name__ == "__main__":
    main()
