"""
12주차 xlsx 파일 단위 비식별화 Apply 테스트 스크립트

실행:
    python notebooks/13_test_xlsx_deidentify_apply.py
"""

from __future__ import annotations

from pathlib import Path
import sys
import unicodedata

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SRC_DIR = PROJECT_ROOT / "src"

if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from deidentify_target_builder import DeidentifyPlan, DeidentifyTarget
from xlsx_deidentify_apply import apply_plan_to_xlsx


OUTPUT_DIR = PROJECT_ROOT / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)


def make_target(
    label: str,
    matched: str,
    sheet_name,
    cell_ref,
    start,
    end,
    *,
    action: str = "마스킹",
    context: str = "",
    location_label: str | None = None,
    file_type: str = "xlsx",
) -> DeidentifyTarget:
    if location_label is None:
        location_label = f"{sheet_name} 탭 {cell_ref} 셀"

    location_meta = {
        "fileType": file_type,
    }

    if sheet_name is not None:
        location_meta["sheetName"] = sheet_name

    if cell_ref is not None:
        location_meta["cellRef"] = cell_ref

    return DeidentifyTarget(
        label=label,
        matched=matched,
        action=action,
        location_label=location_label,
        location_meta=location_meta,
        start=start,
        end=end,
        source="regex",
        reason="xlsx apply 테스트 target",
        grade="S",
        sensitive_type="개인정보",
        sensitive_category=label,
        context=context,
        order=0,
    )


def create_sample_xlsx(path: Path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "민원처리"

    ws["A1"] = "담당자 이메일은 test@example.com입니다."
    ws["A2"] = "직원 김도윤의 서류를 검토했습니다."
    ws["A3"] = "담당자 김도윤의 이메일은 test@example.com입니다."
    ws["A4"] = "서버 IP는 192.168.0.1입니다."
    ws["A5"] = "입찰 제안 평가표를 검토했습니다."
    ws["A6"] = "담당자 이메일: test@example.com"
    ws["A7"] = "담당자 이메일은 test@example.com입니다."
    ws["A8"] = 1012345678
    ws["A9"] = "=SUM(1,2)"
    ws["B10"] = "병합셀 김도윤"
    ws.merge_cells("B10:D10")
    ws["A11"] = "첫 번째 셀 test@example.com"
    ws["B11"] = "두 번째 셀 김도윤"

    # NFC/NFD 시트명 테스트
    nfd_sheet_name = unicodedata.normalize("NFD", "계약내역")
    ws_nfd = wb.create_sheet(nfd_sheet_name)
    ws_nfd["A1"] = "계약 담당자 김도윤"

    wb.save(path)


def print_common_result(title: str, result) -> None:
    print(f"\n=== {title} ===")
    print(f"fileType: {result.fileType}")
    print(f"inputFilePath: {result.inputFilePath}")
    print(f"outputFilePath: {result.outputFilePath}")

    print("summary:")
    print(f"  totalLocations: {result.summary.totalLocations}")
    print(f"  appliedLocations: {result.summary.appliedLocations}")
    print(f"  partialLocations: {result.summary.partialLocations}")
    print(f"  skippedLocations: {result.summary.skippedLocations}")
    print(f"  totalWarnings: {result.summary.totalWarnings}")
    print(f"  autoTargetCount: {result.summary.autoTargetCount}")
    print(f"  reviewTargetCount: {result.summary.reviewTargetCount}")

    print(f"autoResults: {len(result.autoResults)}")
    for item in result.autoResults:
        print(
            f"  - {item.locationLabel} / {item.label} / status={item.status} / "
            f"applied={item.appliedTargetCount}, skipped={item.skippedTargetCount}"
        )
        print(f"    original: {item.originalText}")
        print(f"    applied : {item.appliedText}")
        for warning in item.warnings:
            print(f"    warning: {warning}")

    print(f"reviewTargets: {len(result.reviewTargets)}")
    for review in result.reviewTargets:
        print(f"  - {review.locationLabel} / {review.label} / action={review.action}")

    print(f"global warnings: {len(result.warnings)}")
    for warning in result.warnings:
        print(f"  - {warning}")


def main() -> None:
    print("=== 12주차 xlsx 파일 단위 비식별화 Apply 테스트 ===")

    input_path = OUTPUT_DIR / "week12_sample.xlsx"
    create_sample_xlsx(input_path)

    targets = [
        # TC1
        make_target(
            "이메일 주소",
            "test@example.com",
            "민원처리",
            "A1",
            9,
            25,
            context="담당자 이메일은 test@example.com입니다.",
        ),
        # TC2
        make_target(
            "성명",
            "김도윤",
            "민원처리",
            "A2",
            3,
            6,
            context="직원 김도윤의 서류를 검토했습니다.",
        ),
        # TC3
        make_target(
            "성명",
            "김도윤",
            "민원처리",
            "A3",
            4,
            7,
            context="담당자 김도윤의 이메일은 test@example.com입니다.",
        ),
        make_target(
            "이메일 주소",
            "test@example.com",
            "민원처리",
            "A3",
            14,
            30,
            context="담당자 김도윤의 이메일은 test@example.com입니다.",
        ),
        # TC4
        make_target(
            "내부 IP 주소",
            "192.168.0.1",
            "민원처리",
            "A4",
            7,
            18,
            action="삭제",
            context="서버 IP는 192.168.0.1입니다.",
        ),
        # TC6 sheetName 없음
        make_target(
            "이메일 주소",
            "test@example.com",
            None,
            "A1",
            9,
            25,
            context="담당자 이메일은 test@example.com입니다.",
            location_label="sheetName 없음 테스트",
        ),
        # TC7 cellRef 없음
        make_target(
            "이메일 주소",
            "test@example.com",
            "민원처리",
            None,
            9,
            25,
            context="담당자 이메일은 test@example.com입니다.",
            location_label="cellRef 없음 테스트",
        ),
        # TC8 context 불일치, slice 일치
        make_target(
            "이메일 주소",
            "test@example.com",
            "민원처리",
            "A6",
            9,
            25,
            context="담당자 이메일은 test@example.com입니다.",
        ),
        # TC9 slice 불일치
        make_target(
            "이메일 주소",
            "test@example.com",
            "민원처리",
            "A7",
            9,
            24,
            context="담당자 이메일은 test@example.com입니다.",
        ),
        # TC10 숫자 셀
        make_target(
            "전화번호",
            "1012345678",
            "민원처리",
            "A8",
            0,
            10,
            context="1012345678",
        ),
        # TC11 수식 셀
        make_target(
            "수식",
            "=SUM(1,2)",
            "민원처리",
            "A9",
            0,
            9,
            context="=SUM(1,2)",
        ),
        # TC12 병합 셀 비좌상단
        make_target(
            "성명",
            "김도윤",
            "민원처리",
            "C10",
            4,
            7,
            context="병합셀 김도윤",
        ),
        # TC13 한글 시트명 NFC 정규화
        make_target(
            "성명",
            "김도윤",
            "계약내역",
            "A1",
            7,
            10,
            context="계약 담당자 김도윤",
        ),
        # TC15 같은 시트 여러 셀
        make_target(
            "이메일 주소",
            "test@example.com",
            "민원처리",
            "A11",
            7,
            23,
            context="첫 번째 셀 test@example.com",
        ),
        make_target(
            "성명",
            "김도윤",
            "민원처리",
            "B11",
            7,
            10,
            context="두 번째 셀 김도윤",
        ),
    ]

    review_targets = [
        make_target(
            "민감정보",
            "",
            "민원처리",
            "A5",
            None,
            None,
            action="검토 필요",
            context="입찰 제안 평가표를 검토했습니다.",
        )
    ]

    plan = DeidentifyPlan(
        auto_targets=targets,
        review_targets=review_targets,
        summary_grade="S",
    )

    output_path = OUTPUT_DIR / "week12_sample_deidentified.xlsx"
    if output_path.exists():
        output_path.unlink()

    result = apply_plan_to_xlsx(
        str(input_path),
        plan,
        output_path=str(output_path),
        deletion_mode="delete",
    )

    print_common_result("xlsx Apply 결과", result)

    # Output workbook 주요 셀 검증 출력
    wb = load_workbook(output_path, data_only=False)
    ws = wb["민원처리"]
    print("\n=== output xlsx 주요 셀 값 확인 ===")
    for cell_ref in ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9", "A11", "B11"]:
        print(f"{cell_ref}: {ws[cell_ref].value}")

    normalized_contract = None
    for sheet_name in wb.sheetnames:
        if unicodedata.normalize("NFC", sheet_name) == unicodedata.normalize("NFC", "계약내역"):
            normalized_contract = sheet_name
            break
    if normalized_contract:
        print(f"계약내역!A1: {wb[normalized_contract]['A1'].value}")

    # TC14 원본 유지 확인
    original_wb = load_workbook(input_path, data_only=False)
    print("\n=== TC14 원본 파일 유지 확인 ===")
    print(f"원본 A1: {original_wb['민원처리']['A1'].value}")
    print(f"결과 A1: {ws['A1'].value}")

    # TC16 summary 정합성 확인
    print("\n=== TC16 summary 정합성 확인 ===")
    total = result.summary.totalLocations
    status_sum = (
        result.summary.appliedLocations
        + result.summary.partialLocations
        + result.summary.skippedLocations
    )
    print(f"totalLocations == len(autoResults): {total == len(result.autoResults)}")
    print(f"status 합계 == totalLocations: {status_sum == total}")
    print(f"status 합계: {status_sum}, totalLocations: {total}")

    print("\n=== 테스트 완료 ===")


if __name__ == "__main__":
    main()
