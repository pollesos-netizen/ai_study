"""
12주차 xlsx Apply 회귀 테스트

13주차 리팩토링 후 xlsx 모듈이 정상 동작하는지 확인합니다.
- common_apply_utils 사용
- applyMode="applied" 명시
- 빈 셀 처리 추가
- cell.data_type="s" 명시
- 코드화된 warning type
"""

from __future__ import annotations

from pathlib import Path
import sys
import tempfile

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from openpyxl import Workbook, load_workbook

from common_apply_result import APPLY_MODE_APPLIED
from deidentify_target_builder import DeidentifyPlan, DeidentifyTarget
from xlsx_deidentify_apply import apply_plan_to_xlsx


_results = []


def _check(tc_id, condition, message=""):
    _results.append((tc_id, condition, message))
    status = "PASS" if condition else "FAIL"
    print(f"  [{status}] {tc_id}{': ' + message if message and not condition else ''}")


def _make_target(label, matched, start, end, source, action, sheet, cell_ref, context, grade="S"):
    return DeidentifyTarget(
        label=label, matched=matched, action=action,
        location_label=f"{sheet} 탭 {cell_ref} 셀",
        location_meta={"fileType": "xlsx", "sheetName": sheet, "cellRef": cell_ref},
        start=start, end=end, source=source, reason="test",
        grade=grade, context=context,
    )


def _create_xlsx(cells, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for cell_ref, value in cells.items():
        ws[cell_ref] = value
    wb.save(str(path))
    return path


def main():
    with tempfile.TemporaryDirectory() as tmp_dir_str:
        tmp_dir = Path(tmp_dir_str)
        print("=== xlsx 회귀 테스트 ===")

        # 1. 기본 이메일 마스킹
        print("\n[기본 이메일 마스킹]")
        path = _create_xlsx({"B12": "담당자 이메일은 test@example.com입니다."}, tmp_dir / "x1.xlsx")
        plan = DeidentifyPlan(
            auto_targets=[_make_target("이메일", "test@example.com", 9, 25,
                                        "regex", "마스킹", "Sheet1", "B12",
                                        "담당자 이메일은 test@example.com입니다.")],
            review_targets=[],
        )
        result = apply_plan_to_xlsx(str(path), plan)
        _check("xlsx.applyMode", result.applyMode == APPLY_MODE_APPLIED)
        _check("xlsx.outputFilePath_exists", result.outputFilePath is not None)
        _check("xlsx.output_file_created", Path(result.outputFilePath).exists())
        _check("xlsx.autoResults", len(result.autoResults) == 1)
        item = result.autoResults[0]
        _check("xlsx.status", item.status == "applied")
        _check("xlsx.applied_count", item.appliedTargetCount == 1)
        _check("xlsx.appliedText_masked",
               "*" * len("test@example.com") in item.appliedText)

        # 2. 출력 파일에서 cell.data_type 확인
        out_wb = load_workbook(result.outputFilePath, data_only=False)
        out_cell = out_wb["Sheet1"]["B12"]
        _check("xlsx.cell_data_type_s", out_cell.data_type == "s",
               f"data_type={out_cell.data_type}")

        # 3. 빈 문자열 셀 처리
        print("\n[빈 문자열 셀 skip + warning]")
        path = _create_xlsx({"A1": ""}, tmp_dir / "x2.xlsx")
        plan = DeidentifyPlan(
            auto_targets=[_make_target("이메일", "test", 0, 4,
                                        "regex", "마스킹", "Sheet1", "A1", "test")],
            review_targets=[],
        )
        result = apply_plan_to_xlsx(str(path), plan)
        item = result.autoResults[0]
        _check("xlsx_empty.status", item.status == "skipped")
        _check("xlsx_empty.warning_type",
               any("[empty_cell]" in w for w in item.warnings),
               f"warnings={item.warnings}")

        # 4. 숫자 셀 처리
        print("\n[숫자 셀 skip + warning]")
        path = _create_xlsx({"A1": 123456789}, tmp_dir / "x3.xlsx")
        plan = DeidentifyPlan(
            auto_targets=[_make_target("전화번호", "123", 0, 3,
                                        "regex", "마스킹", "Sheet1", "A1", "123456789")],
            review_targets=[],
        )
        result = apply_plan_to_xlsx(str(path), plan)
        item = result.autoResults[0]
        _check("xlsx_num.status", item.status == "skipped")
        _check("xlsx_num.warning_type",
               any("[non_string_cell]" in w for w in item.warnings),
               f"warnings={item.warnings}")

        # 5. sheetName 없음
        print("\n[sheetName 누락]")
        path = _create_xlsx({"A1": "test"}, tmp_dir / "x4.xlsx")
        target = DeidentifyTarget(
            label="X", matched="test", action="마스킹",
            location_label="알 수 없음",
            location_meta={"fileType": "xlsx", "cellRef": "A1"},  # sheetName 없음
            start=0, end=4, source="regex", reason="test",
            grade="S", context="test",
        )
        plan = DeidentifyPlan(auto_targets=[target], review_targets=[])
        result = apply_plan_to_xlsx(str(path), plan)
        item = result.autoResults[0]
        _check("xlsx_no_sheet.warning_type",
               any("[missing_sheet_name]" in w for w in item.warnings),
               f"warnings={item.warnings}")

    print("\n=== 결과 요약 ===")
    total = len(_results)
    passed = sum(1 for _, ok, _ in _results if ok)
    print(f"  통과: {passed} / 전체: {total}")
    failed = total - passed
    if failed:
        for tc_id, ok, msg in _results:
            if not ok:
                print(f"    - {tc_id}: {msg}")
        sys.exit(1)


if __name__ == "__main__":
    main()
